from flask import Blueprint, render_template, request, redirect, url_for, session, flash, send_file
from datetime import datetime
from zipfile import ZipFile
import os
import sqlite3
from werkzeug.security import generate_password_hash
from db import get_db
from utils import log_action, login_required
from gen_docx import gen_doc
import logging
import openpyxl
from werkzeug.utils import secure_filename
import pandas as pd

admin_bp = Blueprint('admin', __name__)

@admin_bp.route('/admin/manage_education_documents', methods=['GET', 'POST'])
@login_required('admin')
def manage_education_documents():
    db = get_db()
    cursor = db.cursor()
    
    # Получение списка студентов
    cursor.execute("SELECT id, last_name_UA, first_name_UA FROM students")
    students = cursor.fetchall()

    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'delete':
            doc_id = request.form.get('doc_id')
            try:
                cursor.execute("DELETE FROM education_documents WHERE id = ?", (doc_id,))
                cursor.execute("DELETE FROM foreign_education_docs WHERE education_doc_id = ?", (doc_id,))
                db.commit()
                flash('Документ успішно видалено!', 'success')
            except sqlite3.Error as e:
                db.rollback()
                flash(f'Помилка при видаленні документа: {e}', 'danger')
        elif action == 'edit':
            doc_id = request.form.get('doc_id')
            student_id = request.form.get('student_id')  # Получаем новый student_id
            document_type = request.form.get('document_type')
            document_type_en = request.form.get('document_type_en')
            document_number = request.form.get('document_number')
            institution_name = request.form.get('institution_name')
            institution_name_en = request.form.get('institution_name_en')
            country = request.form.get('country')
            country_en = request.form.get('country_en')
            completion_date = request.form.get('completion_date')
            reference_number = request.form.get('reference_number')
            reference_institution = request.form.get('reference_institution')
            reference_institution_en = request.form.get('reference_institution_en')
            reference_country = request.form.get('reference_country')
            reference_country_en = request.form.get('reference_country_en')
            reference_issue_date = request.form.get('reference_issue_date')
            recognition_certificate_number = request.form.get('recognition_certificate_number')
            recognition_issuer = request.form.get('recognition_issuer')
            recognition_issuer_en = request.form.get('recognition_issuer_en')
            recognition_date = request.form.get('recognition_date')

            try:
                # Проверка существования student_id
                if student_id:
                    cursor.execute("SELECT id FROM students WHERE id = ?", (student_id,))
                    if not cursor.fetchone():
                        flash('Обраний студент не існує!', 'danger')
                        return redirect(url_for('admin.manage_education_documents'))
                else:
                    flash('Не вказано студента!', 'danger')
                    return redirect(url_for('admin.manage_education_documents'))

                cursor.execute(
                    "UPDATE education_documents SET student_id = ?, document_type = ?, document_type_en = ?, document_number = ?, institution_name = ?, institution_name_en = ?, country = ?, country_en = ?, completion_date = ? WHERE id = ?",
                    (student_id, document_type, document_type_en, document_number, institution_name, institution_name_en, country, country_en, completion_date, doc_id)
                )
                cursor.execute("SELECT id FROM foreign_education_docs WHERE education_doc_id = ?", (doc_id,))
                foreign_doc = cursor.fetchone()
                if foreign_doc:
                    cursor.execute(
                        "UPDATE foreign_education_docs SET reference_number = ?, reference_institution = ?, reference_institution_en = ?, reference_country = ?, reference_country_en = ?, reference_issue_date = ?, recognition_certificate_number = ?, recognition_issuer = ?, recognition_issuer_en = ?, recognition_date = ? WHERE education_doc_id = ?",
                        (reference_number, reference_institution, reference_institution_en, reference_country, reference_country_en, reference_issue_date, recognition_certificate_number, recognition_issuer, recognition_issuer_en, recognition_date, doc_id)
                    )
                elif any([reference_number, reference_institution, reference_institution_en, reference_country, reference_country_en, reference_issue_date, recognition_certificate_number, recognition_issuer, recognition_issuer_en, recognition_date]):
                    cursor.execute(
                        "INSERT INTO foreign_education_docs (education_doc_id, reference_number, reference_institution, reference_institution_en, reference_country, reference_country_en, reference_issue_date, recognition_certificate_number, recognition_issuer, recognition_issuer_en, recognition_date) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        (doc_id, reference_number, reference_institution, reference_institution_en, reference_country, reference_country_en, reference_issue_date, recognition_certificate_number, recognition_issuer, recognition_issuer_en, recognition_date)
                    )
                db.commit()
                flash('Документ успішно відредаговано!', 'success')
            except sqlite3.Error as e:
                db.rollback()
                flash(f'Помилка при редагуванні документа: {e}', 'danger')
        else:
            student_id = request.form.get('student_id')
            country = request.form.get('country')
            country_en = request.form.get('country_en')
            document_type = request.form.get('document_type')
            document_type_en = request.form.get('document_type_en')
            document_number = request.form.get('document_number')
            institution_name = request.form.get('institution_name')
            institution_name_en = request.form.get('institution_name_en')
            completion_date = request.form.get('completion_date')

            try:
                cursor.execute(
                    "INSERT INTO education_documents (student_id, document_type, document_type_en, document_number, institution_name, institution_name_en, country, country_en, completion_date) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (student_id, document_type, document_type_en, document_number, institution_name, institution_name_en, country, country_en, completion_date)
                )
                education_doc_id = cursor.lastrowid

                if country.lower() != 'україна' and country.lower() != 'ukraine':
                    reference_number = request.form.get('reference_number')
                    reference_institution = request.form.get('reference_institution')
                    reference_institution_en = request.form.get('reference_institution_en')
                    reference_country = request.form.get('reference_country')
                    reference_country_en = request.form.get('reference_country_en')
                    reference_issue_date = request.form.get('reference_issue_date')
                    recognition_certificate_number = request.form.get('recognition_certificate_number')
                    recognition_issuer = request.form.get('recognition_issuer')
                    recognition_issuer_en = request.form.get('recognition_issuer_en')
                    recognition_date = request.form.get('recognition_date')

                    if any([reference_number, reference_institution, reference_institution_en, reference_country, reference_country_en, reference_issue_date, recognition_certificate_number, recognition_issuer, recognition_issuer_en, recognition_date]):
                        cursor.execute(
                            "INSERT INTO foreign_education_docs (education_doc_id, reference_number, reference_institution, reference_institution_en, reference_country, reference_country_en, reference_issue_date, recognition_certificate_number, recognition_issuer, recognition_issuer_en, recognition_date) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                            (education_doc_id, reference_number, reference_institution, reference_institution_en, reference_country, reference_country_en, reference_issue_date, recognition_certificate_number, recognition_issuer, recognition_issuer_en, recognition_date)
                        )

                db.commit()
                flash('Документ успішно додано!', 'success')
            except sqlite3.Error as e:
                db.rollback()
                flash(f'Помилка при додаванні документа: {e}', 'danger')
            finally:
                cursor.close()
                return redirect(url_for('admin.manage_education_documents'))

    # Получение списка документов с проверкой существования student_id
    cursor.execute("""
        SELECT ed.id, s.archived, s.last_name_UA, s.first_name_UA, ed.student_id, ed.document_type, ed.document_number, ed.document_type_en, ed.institution_name, ed.institution_name_en, ed.country, ed.country_en, ed.completion_date,
               fed.reference_number, fed.reference_institution, fed.reference_institution_en, fed.reference_country, fed.reference_country_en, fed.reference_issue_date, fed.recognition_certificate_number, fed.recognition_issuer, fed.recognition_issuer_en, fed.recognition_date
        FROM education_documents ed
        LEFT JOIN students s ON ed.student_id = s.id
        LEFT JOIN foreign_education_docs fed ON ed.id = fed.education_doc_id
        WHERE ed.student_id IS NOT NULL AND ed.student_id IN (SELECT id FROM students) AND s.archived = FALSE
    """)
    documents = cursor.fetchall()

    cursor.close()
    return render_template('manage_education_documents.html', students=students, documents=documents)

@admin_bp.route('/admin/groups', methods=['GET', 'POST'])
@login_required('admin')
def manage_groups():
    """Управление группами: просмотр, добавление, редактирование и удаление групп."""
    conn = get_db()
    conn.row_factory = sqlite3.Row
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'add':
            # Добавление новой группы
            name = request.form.get('name')
            start_year = request.form.get('start_year')
            study_form = request.form.get('study_form')
            program_credits = request.form.get('program_credits')
            qualification_name = request.form.get('qualification_name')
            degree_level = request.form.get('degree_level')
            specialty = request.form.get('specialty')
            educational_program = request.form.get('educational_program')
            knowledge_area = request.form.get('knowledge_area')
            qualification_name_en = request.form.get('qualification_name_en')
            degree_level_en = request.form.get('degree_level_en')
            specialty_en = request.form.get('specialty_en')
            educational_program_en = request.form.get('educational_program_en')
            knowledge_area_en = request.form.get('knowledge_area_en')
            institution_name_and_status = request.form.get('institution_name_and_status')
            institution_name_and_status_en = request.form.get('institution_name_and_status_en')
            entry_requirements = request.form.get('entry_requirements')
            entry_requirements_en = request.form.get('entry_requirements_en')
            learning_outcomes = request.form.get('learning_outcomes')
            learning_outcomes_en = request.form.get('learning_outcomes_en')
            program_includes = request.form.get('program_includes')
            program_includes_en = request.form.get('program_includes_en')

            # Валидация данных
            required_fields = [
                name, start_year, study_form, program_credits,
                # qualification_name, degree_level, specialty, educational_program, knowledge_area,
                # qualification_name_en, degree_level_en, specialty_en, educational_program_en, knowledge_area_en,
                # institution_name_and_status, institution_name_and_status_en,
                # entry_requirements, entry_requirements_en,
                # learning_outcomes, learning_outcomes_en, program_includes, program_includes_en
            ]
            if not all(required_fields):
                flash("Усі поля мають бути заповнені.", "error")
            elif study_form not in ['Денна', 'Заочна']:
                flash("Форма навчання має бути 'Денна' або 'Заочна'.", "error")
            elif program_credits not in ['180', '240']:
                flash("Кількість кредитів має бути 180 або 240.", "error")
            else:
                try:
                    start_year = int(start_year)
                    program_credits = int(program_credits)
                    if start_year < 2000 or start_year > 2025:
                        flash("Рік початку навчання має бути між 2000 і 2025.", "error")
                    else:
                        conn.execute("""
                            INSERT INTO groups (
                                name, start_year, study_form, program_credits,
                                qualification_name, degree_level, specialty, educational_program, knowledge_area,
                                qualification_name_en, degree_level_en, specialty_en, educational_program_en, knowledge_area_en,
                                institution_name_and_status, institution_name_and_status_en,
                                entry_requirements, entry_requirements_en,
                                learning_outcomes, learning_outcomes_en, program_includes, program_includes_en
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (
                            name, start_year, study_form, program_credits,
                            qualification_name, degree_level, specialty, educational_program, knowledge_area,
                            qualification_name_en, degree_level_en, specialty_en, educational_program_en, knowledge_area_en,
                            institution_name_and_status, institution_name_and_status_en,
                            entry_requirements, entry_requirements_en,
                            learning_outcomes, learning_outcomes_en, program_includes, program_includes_en
                        ))
                        conn.commit()
                        flash("Групу додано успішно.", "success")
                        log_action(
                            session.get('username', 'невідомо'),
                            f"додав групу {name} ({start_year}, {study_form}, {program_credits} кредитів, {qualification_name})",
                            [conn.execute("SELECT last_insert_rowid()").fetchone()[0]]
                        )
                except ValueError:
                    flash("Рік початку навчання або кредити мають бути числами.", "error")
                except sqlite3.IntegrityError:
                    flash("Група з таким назвою та роком початку навчання вже існує.", "error")

        elif action == 'edit':
            # Редактирование группы
            group_id = request.form.get('group_id')
            name = request.form.get('name')
            start_year = request.form.get('start_year')
            study_form = request.form.get('study_form')
            program_credits = request.form.get('program_credits')
            qualification_name = request.form.get('qualification_name')
            degree_level = request.form.get('degree_level')
            specialty = request.form.get('specialty')
            educational_program = request.form.get('educational_program')
            knowledge_area = request.form.get('knowledge_area')
            qualification_name_en = request.form.get('qualification_name_en')
            degree_level_en = request.form.get('degree_level_en')
            specialty_en = request.form.get('specialty_en')
            educational_program_en = request.form.get('educational_program_en')
            knowledge_area_en = request.form.get('knowledge_area_en')
            institution_name_and_status = request.form.get('institution_name_and_status')
            institution_name_and_status_en = request.form.get('institution_name_and_status_en')
            entry_requirements = request.form.get('entry_requirements')
            entry_requirements_en = request.form.get('entry_requirements_en')
            learning_outcomes = request.form.get('learning_outcomes')
            learning_outcomes_en = request.form.get('learning_outcomes_en')
            program_includes = request.form.get('program_includes')
            program_includes_en = request.form.get('program_includes_en')

            # Валидация данных
            required_fields = [
                group_id, name, start_year, study_form, program_credits,
                # qualification_name, degree_level, specialty, educational_program, knowledge_area,
                # qualification_name_en, degree_level_en, specialty_en, educational_program_en, knowledge_area_en,
                # institution_name_and_status, institution_name_and_status_en,
                # entry_requirements, entry_requirements_en,
                # learning_outcomes, learning_outcomes_en, program_includes, program_includes_en
            ]
            if not all(required_fields):
                flash("Усі поля мають бути заповнені.", "error")
            elif study_form not in ['Денна', 'Заочна']:
                flash("Форма навчання має бути 'Денна' або 'Заочна'.", "error")
            elif program_credits not in ['180', '240']:
                flash("Кількість кредитів має бути 180 або 240.", "error")
            else:
                try:
                    start_year = int(start_year)
                    program_credits = int(program_credits)
                    if start_year < 2000 or start_year > 2025:
                        flash("Рік початку навчання має бути між 2000 і 2025.", "error")
                    else:
                        conn.execute("""
                            UPDATE groups
                            SET name = ?, start_year = ?, study_form = ?, program_credits = ?,
                                qualification_name = ?, degree_level = ?, specialty = ?,
                                educational_program = ?, knowledge_area = ?,
                                qualification_name_en = ?, degree_level_en = ?, specialty_en = ?,
                                educational_program_en = ?, knowledge_area_en = ?,
                                institution_name_and_status = ?, institution_name_and_status_en = ?,
                                entry_requirements = ?, entry_requirements_en = ?,
                                learning_outcomes = ?, learning_outcomes_en = ?,
                                program_includes = ?, program_includes_en = ?
                            WHERE id = ?
                        """, (
                            name, start_year, study_form, program_credits,
                            qualification_name, degree_level, specialty, educational_program, knowledge_area,
                            qualification_name_en, degree_level_en, specialty_en, educational_program_en, knowledge_area_en,
                            institution_name_and_status, institution_name_and_status_en,
                            entry_requirements, entry_requirements_en,
                            learning_outcomes, learning_outcomes_en, program_includes, program_includes_en,
                            group_id
                        ))
                        conn.commit()
                        flash("Групу відредаговано успішно.", "success")
                        log_action(
                            session.get('username', 'невідомо'),
                            f"відредагував групу ID {group_id} на {name} ({start_year}, {study_form}, {program_credits} кредитів, {qualification_name})",
                            [int(group_id)]
                        )
                except ValueError:
                    flash("Рік початку навчання або кредити мають бути числами.", "error")
                except sqlite3.IntegrityError:
                    flash("Група з таким назвою та роком початку навчання вже існує.", "error")

        elif action == 'delete':
            # Удаление группы
            group_id = request.form.get('group_id')
            # Проверка наличия связанных данных
            related_data = conn.execute("""
                SELECT (SELECT COUNT(*) FROM students WHERE group_id = ?) +
                       (SELECT COUNT(*) FROM subjects WHERE group_id = ?) +
                       (SELECT COUNT(*) FROM practices WHERE group_id = ?) +
                       (SELECT COUNT(*) FROM courseworks WHERE group_id = ?) +
                       (SELECT COUNT(*) FROM attestations WHERE group_id = ?) AS total
            """, (group_id, group_id, group_id, group_id, group_id)).fetchone()['total']
            
            if related_data > 0:
                flash("Неможливо видалити групу, оскільки вона має пов'язані дані (студенти, предмети, практики, курсові роботи або атестації).", "error")
            else:
                conn.execute("DELETE FROM groups WHERE id = ?", (group_id,))
                conn.commit()
                flash("Групу видалено успішно.", "success")
                log_action(
                    session.get('username', 'невідомо'),
                    f"видалив групу ID {group_id}",
                    [int(group_id)]
                )

    # Получение списка групп с количеством студентов
    groups = conn.execute("""
    SELECT g.id, g.name, g.start_year, g.study_form, g.program_credits,
           g.qualification_name, g.degree_level, g.specialty, g.educational_program, g.knowledge_area,
           g.qualification_name_en, g.degree_level_en, g.specialty_en, g.educational_program_en, g.knowledge_area_en,
           g.institution_name_and_status, g.institution_name_and_status_en,
           g.entry_requirements, g.entry_requirements_en,
           g.learning_outcomes, g.learning_outcomes_en, g.program_includes, g.program_includes_en,
           g.name || ' (' || g.start_year || ', ' || g.study_form || ', ' || g.program_credits || ' кредитів)' AS display_name,
           (SELECT COUNT(*) FROM students s WHERE s.group_id = g.id) AS student_count
    FROM groups g
    WHERE g.archived = FALSE
    ORDER BY g.id, g.start_year
    """).fetchall()
    
    conn.close()
    return render_template("manage_groups.html", groups=groups)
  
@admin_bp.route('/admin/subjects', methods=['GET', 'POST'])
@login_required('admin')
def manage_subjects():
    """Управление предметами."""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT id, name, start_year, study_form, program_credits,
               name || ' (' || start_year || ', ' || study_form || ', ' || program_credits || ' кредитів)' AS display_name
        FROM groups
        WHERE archived = FALSE
        ORDER BY name, start_year
    """)
    groups = cursor.fetchall()
    
    selected_group_id = request.args.get('group_id')
    subjects = []
    students = []
    grades = []
    selected_subject_id = request.args.get('subject_id')
    
    if selected_group_id:
        cursor.execute('SELECT * FROM subjects WHERE group_id = ? ORDER BY position', (selected_group_id,))
        subjects = cursor.fetchall()
        if selected_subject_id:
            cursor.execute('SELECT * FROM students WHERE group_id = ? ORDER BY last_name_UA', (selected_group_id,))
            students = cursor.fetchall()
            cursor.execute('SELECT id, student_id, subject_id, grade FROM grades WHERE subject_id = ?', (selected_subject_id,))
            grades = cursor.fetchall()
    
    if request.method == 'POST':
        action = request.form['action']
        group_id = request.form['group_id']
        
        if action == 'add':
            try:
                code = request.form['code'].strip()
                name = request.form['name'].strip()
                credits = int(request.form['credits'])
                type_ = request.form['type']
                position = int(request.form['position'])
                if not code or not name or credits < 1 or position < 1 or type_ not in ['Залік', 'Екзамен']:
                    flash('Некорректные данные предмета', 'error')
                else:
                    cursor.execute('SELECT MAX(position) FROM subjects WHERE group_id = ?', (group_id,))
                    max_position = cursor.fetchone()[0] or 0
                    if position <= max_position:
                        cursor.execute('UPDATE subjects SET position = position + 1 WHERE position >= ? AND group_id = ?', (position, group_id))
                    cursor.execute(
                        'INSERT INTO subjects (code, name, credits, type, position, group_id) VALUES (?, ?, ?, ?, ?, ?)',
                        (code, name, credits, type_, position, group_id)
                    )
                    conn.commit()
                    flash(f'Добавлен предмет {code}', 'success')
            except (KeyError, ValueError):
                flash('Некорректные данные предмета', 'error')
        
        elif action == 'edit':
            try:
                subject_id = request.form['subject_id']
                code = request.form['code'].strip()
                name = request.form['name'].strip()
                credits = int(request.form['credits'])
                type_ = request.form['type']
                position = int(request.form['position'])
                if not code or not name or credits < 1 or position < 1 or type_ not in ['Залік', 'Екзамен']:
                    flash('Некорректные данные предмета', 'error')
                else:
                    cursor.execute('SELECT position FROM subjects WHERE id = ? AND group_id = ?', (subject_id, group_id))
                    current_position = cursor.fetchone()[0]
                    cursor.execute('UPDATE subjects SET position = 0 WHERE id = ? AND group_id = ?', (subject_id, group_id))
                    cursor.execute('UPDATE subjects SET position = position + 1 WHERE position >= ? AND group_id = ? AND id != ?', (position, group_id, subject_id))
                    cursor.execute(
                        'UPDATE subjects SET code = ?, name = ?, credits = ?, type = ?, position = ? WHERE id = ? AND group_id = ?',
                        (code, name, credits, type_, position, subject_id, group_id)
                    )
                    cursor.execute('SELECT id, position FROM subjects WHERE group_id = ? ORDER BY position, id', (group_id,))
                    subjects = cursor.fetchall()
                    for i, subject in enumerate(subjects, 1):
                        if subject['position'] != i:
                            cursor.execute('UPDATE subjects SET position = ? WHERE id = ? AND group_id = ?', (i, subject['id'], group_id))
                    conn.commit()
                    flash(f'Обновлен предмет {code}', 'success')
            except (KeyError, ValueError):
                flash('Некорректные данные предмета', 'error')
        
        elif action == 'delete':
            try:
                subject_id = request.form['subject_id']
                cursor.execute('SELECT position FROM subjects WHERE id = ? AND group_id = ?', (subject_id, group_id))
                position = cursor.fetchone()[0]
                cursor.execute('DELETE FROM subjects WHERE id = ? AND group_id = ?', (subject_id, group_id))
                cursor.execute('UPDATE subjects SET position = position - 1 WHERE position > ? AND group_id = ?', (position, group_id))
                conn.commit()
                flash('Предмет удален', 'success')
            except (KeyError, ValueError):
                flash('Ошибка при удалении предмета', 'error')
        
        elif action == 'move_up':
            try:
                subject_id = request.form['subject_id']
                cursor.execute('SELECT position FROM subjects WHERE id = ? AND group_id = ?', (subject_id, group_id))
                current_position = cursor.fetchone()[0]
                cursor.execute('SELECT id, position FROM subjects WHERE position < ? AND group_id = ? ORDER BY position DESC LIMIT 1', (current_position, group_id))
                prev_subject = cursor.fetchone()
                if prev_subject:
                    prev_id, prev_position = prev_subject['id'], prev_subject['position']
                    cursor.execute('UPDATE subjects SET position = ? WHERE id = ? AND group_id = ?', (prev_position, subject_id, group_id))
                    cursor.execute('UPDATE subjects SET position = ? WHERE id = ? AND group_id = ?', (current_position, prev_id, group_id))
                    conn.commit()
                    flash('Предмет перемещен вверх', 'success')
            except (KeyError, ValueError):
                flash('Ошибка при перемещении предмета', 'error')
        
        elif action == 'move_down':
            try:
                subject_id = request.form['subject_id']
                cursor.execute('SELECT position FROM subjects WHERE id = ? AND group_id = ?', (subject_id, group_id))
                current_position = cursor.fetchone()[0]
                cursor.execute('SELECT id, position FROM subjects WHERE position > ? AND group_id = ? ORDER BY position ASC LIMIT 1', (current_position, group_id))
                next_subject = cursor.fetchone()
                if next_subject:
                    next_id, next_position = next_subject['id'], next_subject['position']
                    cursor.execute('UPDATE subjects SET position = ? WHERE id = ? AND group_id = ?', (next_position, subject_id, group_id))
                    cursor.execute('UPDATE subjects SET position = ? WHERE id = ? AND group_id = ?', (current_position, next_id, group_id))
                    conn.commit()
                    flash('Предмет перемещен вниз', 'success')
            except (KeyError, ValueError):
                flash('Ошибка при перемещении предмета', 'error')
        
        elif action == 'edit_grades':
            try:
                subject_id = request.form['subject_id']
                cursor.execute('SELECT id FROM students WHERE group_id = ?', (group_id,))
                student_ids = [row['id'] for row in cursor.fetchall()]
                for student_id in student_ids:
                    grade_key = f'grade_{student_id}'
                    grade_id_key = f'grade_id_{student_id}'
                    grade = request.form.get(grade_key)
                    grade_id = request.form.get(grade_id_key)
                    
                    if grade:
                        try:
                            grade = int(grade)
                            if not (0 <= grade <= 100):
                                flash(f'Оценка для студента {student_id} должна быть от 0 до 100', 'error')
                                continue
                            if grade_id:
                                cursor.execute(
                                    'UPDATE grades SET grade = ? WHERE id = ? AND student_id = ? AND subject_id = ?',
                                    (grade, grade_id, student_id, subject_id)
                                )
                            else:
                                cursor.execute(
                                    'INSERT INTO grades (student_id, subject_id, grade) VALUES (?, ?, ?)',
                                    (student_id, subject_id, grade)
                                )
                        except ValueError:
                            flash(f'Некорректная оценка для студента {student_id}', 'error')
                            continue
                    else:
                        if grade_id:
                            cursor.execute('DELETE FROM grades WHERE id = ? AND student_id = ? AND subject_id = ?', (grade_id, student_id, subject_id))
                conn.commit()
                flash('Оценки обновлены', 'success')
            except (KeyError, ValueError) as e:
                flash(f'Ошибка при обновлении оценок: {str(e)}', 'error')
        
        conn.close()
        return redirect(url_for('admin.manage_subjects', group_id=group_id))
    
    conn.close()
    return render_template('admin_subjects.html', groups=groups, selected_group_id=selected_group_id, 
                         subjects=subjects, students=students, grades=grades, selected_subject_id=selected_subject_id)

@admin_bp.route('/admin/activities', methods=['GET', 'POST'])
@login_required('admin')
def manage_activities():
    """Управление активностями."""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT id, name, start_year, study_form, program_credits,
               name || ' (' || start_year || ', ' || study_form || ', ' || program_credits || ' кредитів)' AS display_name
        FROM groups
        WHERE archived = FALSE
        ORDER BY name, start_year
    """)
    groups = cursor.fetchall()
    
    selected_group_id = request.args.get('group_id')
    selected_entity_type = request.args.get('entity_type', 'practice')
    selected_entity_id = request.args.get('entity_id')
    
    entities = []
    students = []
    grades = []
    
    if selected_group_id:
        try:
            selected_group_id = int(selected_group_id)
            cursor.execute('SELECT id FROM groups WHERE id = ?', (selected_group_id,))
            if not cursor.fetchone():
                flash('Обрана група не існує', 'error')
                selected_group_id = ''
            else:
                entity_table = {
                    'practice': 'practices',
                    'coursework': 'courseworks',
                    'attestation': 'attestations'
                }.get(selected_entity_type, 'practices')
                
                cursor.execute(f'SELECT * FROM {entity_table} WHERE group_id = ? ORDER BY position', (selected_group_id,))
                entities = cursor.fetchall()
                
                if selected_entity_id:
                    try:
                        selected_entity_id = int(selected_entity_id)
                        cursor.execute('SELECT * FROM students WHERE group_id = ? ORDER BY last_name_UA', (selected_group_id,))
                        students = cursor.fetchall()
                        cursor.execute('SELECT id, student_id, entity_id, entity_type, grade, name FROM activity_grades WHERE entity_id = ? AND entity_type = ?', 
                                     (selected_entity_id, selected_entity_type))
                        grades = cursor.fetchall()
                    except ValueError:
                        flash('Некоректний ID діяльності', 'error')
                        selected_entity_id = ''
        except ValueError:
            flash('Некоректний ID групи', 'error')
            selected_group_id = ''
    
    if request.method == 'POST':
        action = request.form.get('action')
        group_id = request.form.get('group_id')
        entity_type = request.form.get('entity_type', 'practice')
        entity_table = {'practice': 'practices', 'coursework': 'courseworks', 'attestation': 'attestations'}.get(entity_type, 'practices')
        
        try:
            if not group_id:
                flash('ID групи не вказано', 'error')
                return redirect(url_for('admin.manage_activities', group_id=selected_group_id, entity_type=entity_type))
            
            group_id = int(group_id)
            cursor.execute('SELECT id FROM groups WHERE id = ?', (group_id,))
            if not cursor.fetchone():
                flash('Обрана група не існує', 'error')
                return redirect(url_for('admin.manage_activities', entity_type=entity_type))
            
            if action == 'add':
                code = request.form.get('code')
                name = request.form.get('name')
                credits = request.form.get('credits')
                type_ = request.form.get('type')
                position = request.form.get('position')
                
                if not all([code, name, credits, type_, position]) and entity_type != 'attestation':
                    flash('Усі поля мають бути заповнені', 'error')
                    return redirect(url_for('admin.manage_activities', group_id=group_id, entity_type=entity_type))
                
                credits = int(credits) if credits else 0
                position = int(position) if position else 1
                
                if type_ not in ['Залік', 'Екзамен']:
                    flash('Невірний тип оцінки', 'error')
                    return redirect(url_for('admin.manage_activities', group_id=group_id, entity_type=entity_type))
                
                cursor.execute(f'SELECT MAX(position) FROM {entity_table} WHERE group_id = ?', (group_id,))
                max_position = cursor.fetchone()[0] or 0
                if position <= max_position:
                    cursor.execute(f'UPDATE {entity_table} SET position = position + 1 WHERE position >= ? AND group_id = ?', 
                                 (position, group_id))
                
                cursor.execute(
                    f'INSERT INTO {entity_table} (code, name, credits, type, position, group_id) VALUES (?, ?, ?, ?, ?, ?)',
                    (code, name or '', credits, type_, position, group_id)
                )
                conn.commit()
                flash('Діяльність додано', 'success')
            
            elif action == 'edit':
                entity_id = request.form.get('entity_id')
                if not entity_id:
                    flash('ID діяльності не вказано', 'error')
                    return redirect(url_for('admin.manage_activities', group_id=group_id, entity_type=entity_type))
                
                entity_id = int(entity_id)
                code = request.form.get('code')
                name = request.form.get('name')
                credits = request.form.get('credits')
                type_ = request.form.get('type')
                position = request.form.get('position')
                
                if not all([code, name, credits, type_, position]) and entity_type != 'attestation':
                    flash('Усі поля мають бути заповнені', 'error')
                    return redirect(url_for('admin.manage_activities', group_id=group_id, entity_type=entity_type))
                
                credits = int(credits) if credits else 0
                position = int(position)
                
                # Проверка существования записи
                cursor.execute(f'SELECT id FROM {entity_table} WHERE id = ? AND group_id = ?', (entity_id, group_id))
                if not cursor.fetchone():
                    flash('Діяльність з вказаним ID не знайдено', 'error')
                    return redirect(url_for('admin.manage_activities', group_id=group_id, entity_type=entity_type))
                
                cursor.execute(f'SELECT position FROM {entity_table} WHERE id = ?', (entity_id,))
                current_position = cursor.fetchone()[0]
                
                cursor.execute(f'SELECT MAX(position) FROM {entity_table} WHERE group_id = ?', (group_id,))
                max_position = cursor.fetchone()[0] or 0
                
                if position != current_position and position <= max_position:
                    cursor.execute(f'UPDATE {entity_table} SET position = position + 1 WHERE position >= ? AND group_id = ? AND id != ?', 
                                 (position, group_id, entity_id))
                
                cursor.execute(
                    f'UPDATE {entity_table} SET code = ?, name = ?, credits = ?, type = ?, position = ? WHERE id = ? AND group_id = ?',
                    (code, name or '', credits, type_, position, entity_id, group_id)
                )
                conn.commit()
                flash('Діяльність оновлено', 'success')
            
            elif action == 'delete':
                entity_id = request.form.get('entity_id')
                entity_id = int(entity_id)
                
                cursor.execute(f'SELECT position FROM {entity_table} WHERE id = ?', (entity_id,))
                position = cursor.fetchone()[0]
                
                cursor.execute(f'DELETE FROM {entity_table} WHERE id = ? AND group_id = ?', (entity_id, group_id))
                cursor.execute(f'UPDATE {entity_table} SET position = position - 1 WHERE position > ? AND group_id = ?', 
                             (position, group_id))
                cursor.execute('DELETE FROM activity_grades WHERE entity_id = ? AND entity_type = ?', 
                             (entity_id, entity_type))
                conn.commit()
                flash('Діяльність видалено', 'success')
            
            elif action == 'move_up':
                entity_id = request.form.get('entity_id')
                entity_id = int(entity_id)
                
                cursor.execute(f'SELECT position FROM {entity_table} WHERE id = ?', (entity_id,))
                current_position = cursor.fetchone()[0]
                
                if current_position > 1:
                    cursor.execute(f'UPDATE {entity_table} SET position = ? WHERE position = ? AND group_id = ?', 
                                 (current_position, current_position - 1, group_id))
                    cursor.execute(f'UPDATE {entity_table} SET position = ? WHERE id = ? AND group_id = ?', 
                                 (current_position - 1, entity_id, group_id))
                    conn.commit()
                    flash('Діяльність переміщено вгору', 'success')
            
            elif action == 'move_down':
                entity_id = request.form.get('entity_id')
                entity_id = int(entity_id)
                
                cursor.execute(f'SELECT position FROM {entity_table} WHERE id = ?', (entity_id,))
                current_position = cursor.fetchone()[0]
                
                cursor.execute(f'SELECT MAX(position) FROM {entity_table} WHERE group_id = ?', (group_id,))
                max_position = cursor.fetchone()[0]
                
                if current_position < max_position:
                    cursor.execute(f'UPDATE {entity_table} SET position = ? WHERE position = ? AND group_id = ?', 
                                 (current_position, current_position + 1, group_id))
                    cursor.execute(f'UPDATE {entity_table} SET position = ? WHERE id = ? AND group_id = ?', 
                                 (current_position + 1, entity_id, group_id))
                    conn.commit()
                    flash('Діяльність переміщено вниз', 'success')
            
            elif action == 'edit_grades':
                try:
                    entity_id = request.form['entity_id']
                    cursor.execute('SELECT id FROM students WHERE group_id = ?', (group_id,))
                    student_ids = [row['id'] for row in cursor.fetchall()]
                    for student_id in student_ids:
                        grade_key = f'grade_{student_id}'
                        grade_id_key = f'grade_id_{student_id}'
                        name_key = f'name_{student_id}'
                        grade = request.form.get(grade_key)
                        grade_id = request.form.get(grade_id_key)
                        name = request.form.get(name_key, '') if entity_type == 'attestation' else ''
                        
                        if grade:
                            try:
                                grade = int(grade)
                                if not (0 <= grade <= 100):
                                    flash(f'Оценка для студента {student_id} должна быть от 0 до 100', 'error')
                                    continue
                                if grade_id:
                                    cursor.execute(
                                        'UPDATE activity_grades SET grade = ?, name = ? WHERE id = ? AND student_id = ? AND entity_id = ? AND entity_type = ?',
                                        (grade, name, grade_id, student_id, entity_id, entity_type)
                                    )
                                else:
                                    cursor.execute(
                                        'INSERT INTO activity_grades (student_id, entity_id, entity_type, grade, name) VALUES (?, ?, ?, ?, ?)',
                                        (student_id, entity_id, entity_type, grade, name)
                                    )
                            except ValueError:
                                flash(f'Некорректная оценка для студента {student_id}', 'error')
                                continue
                        else:
                            if grade_id:
                                cursor.execute(
                                    'DELETE FROM activity_grades WHERE id = ? AND student_id = ? AND entity_id = ? AND entity_type = ?',
                                    (grade_id, student_id, entity_id, entity_type)
                                )
                    conn.commit()
                    flash('Оценки обновлены', 'success')
                except (KeyError, ValueError) as e:
                    flash(f'Ошибка при обновлении оценок: {str(e)}', 'error')
            
            return redirect(url_for('admin.manage_activities', group_id=group_id, entity_type=entity_type))
        
        except (ValueError, sqlite3.Error) as e:
            conn.rollback()
            flash(f'Помилка: {e}', 'error')
            return redirect(url_for('admin.manage_activities', group_id=selected_group_id, entity_type=entity_type))
    
    conn.close()
    return render_template('admin_activities.html', 
                         groups=groups, 
                         selected_group_id=selected_group_id, 
                         entities=entities, 
                         students=students, 
                         grades=grades, 
                         selected_entity_id=selected_entity_id,
                         entity_type=selected_entity_type)
                         
@admin_bp.route('/admin/logs')
@login_required('admin')
def view_logs():
    """Отображение логов действий пользователей."""
    current_dir = os.path.dirname(__file__)
    project_root = os.path.dirname(current_dir)
    log_file_path = os.path.join(project_root, 'app.log')
    
    logs = []
    if os.path.exists(log_file_path):
        try:
            with open(log_file_path, 'r', encoding='utf-8') as file:
                logs = [line.strip() for line in file if line.strip()]
        except Exception as e:
            logging.error(f"Ошибка при чтении файла: {e}")
    
    log_action(session.get('username', 'невідомо'), "переглянув логи дій користувачів")
    return render_template('view_logs.html', logs=logs[::-1])

@admin_bp.route('/admin/users')
@login_required('admin')
def user_list():
    """Отображение списка пользователей."""
    conn = get_db()
    users = conn.execute("""
        SELECT u.id, u.username, u.role,
               GROUP_CONCAT(
                   g.name || ' (' || g.start_year || ', ' || g.study_form || ', ' || g.program_credits || ' кредитів)',
                   ', '
               ) AS group_names
        FROM users u
        LEFT JOIN user_groups ug ON u.id = ug.user_id
        LEFT JOIN groups g ON ug.group_id = g.id
        GROUP BY u.id
    """).fetchall()
    
    log_action(session.get('username', 'невідомо'), "переглянув список користувачів")
    conn.close()
    return render_template('user_list.html', users=users)

@admin_bp.route('/admin/users/add', methods=['GET', 'POST'])
@login_required('admin')
def add_user():
    """Добавление нового пользователя."""
    conn = get_db()

    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        role = request.form['role']
        group_ids = request.form.getlist('group_id')  # Multiple selection
        password_hash = generate_password_hash(password)

        try:
            # Выполняем вставку в таблицу users
            cursor = conn.execute("""
                INSERT INTO users (username, password_hash, role)
                VALUES (?, ?, ?)
            """, (username, password_hash, role))
            user_id = cursor.lastrowid  # Получаем ID нового пользователя

            # Добавляем связи с группами
            for group_id in group_ids:
                if group_id:
                    conn.execute("""
                        INSERT INTO user_groups (user_id, group_id)
                        VALUES (?, ?)
                    """, (user_id, group_id))

            conn.commit()
            log_action(session.get('username', 'невідомо'), f"додав користувача {username}")
            conn.close()
            return redirect(url_for('admin.user_list'))
        except sqlite3.IntegrityError as e:
            conn.rollback()  # Откатываем изменения при ошибке
            flash('Користувач з таким ім\'ям уже існує.', 'error')
            logging.error(f"Помилка додавання користувача {username}: {e}")
        except Exception as e:
            conn.rollback()  # Откатываем изменения при любой другой ошибке
            flash('Сталася помилка при додаванні користувача.', 'error')
            logging.error(f"Помилка додавання користувача {username}: {e}")

    groups = conn.execute("""
        SELECT id, name, start_year, study_form, program_credits,
               name || ' (' || start_year || ', ' || study_form || ', ' || program_credits || ' кредитів)' AS display_name
        FROM groups
        ORDER BY name, start_year
    """).fetchall()
    conn.close()
    log_action(session.get('username', 'невідомо'), "відкрив форму додавання користувача")
    return render_template('add_user.html', groups=groups)

@admin_bp.route('/admin/users/<int:user_id>/delete')
@login_required('admin')
def delete_user(user_id):
    """Удаление пользователя."""
    conn = get_db()
    conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
    conn.commit()
    log_action(session.get('username', 'невідомо'), f"видалив користувача ID {user_id}")
    conn.close()
    return redirect(url_for('admin.user_list'))

@admin_bp.route('/admin/users/<int:user_id>/change_password', methods=['GET', 'POST'])
@login_required('admin')
def change_password(user_id):
    """Смена пароля пользователя."""
    if request.method == 'POST':
        new_password = request.form['password']
        new_hash = generate_password_hash(new_password)
        conn = get_db()
        conn.execute("UPDATE users SET password_hash = ? WHERE id = ?", (new_hash, user_id))
        conn.commit()
        log_action(session.get('username', 'невідомо'), f"змінив пароль користувача ID {user_id}")
        conn.close()
        return redirect(url_for('admin.user_list'))

    log_action(session.get('username', 'невідомо'), f"відкрив форму зміни пароля для користувача ID {user_id}")
    return render_template('change_password.html', user_id=user_id)

@admin_bp.route('/admin/users/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required('admin')
def edit_user(user_id):
    """Редактирование данных пользователя."""
    conn = get_db()

    if request.method == 'POST':
        new_role = request.form['role']
        group_ids = request.form.getlist('group_id')  # Multiple selection
        conn.execute("UPDATE users SET role = ? WHERE id = ?", (new_role, user_id))
        conn.execute("DELETE FROM user_groups WHERE user_id = ?", (user_id,))
        for group_id in group_ids:
            if group_id:
                conn.execute("""
                    INSERT INTO user_groups (user_id, group_id)
                    VALUES (?, ?)
                """, (user_id, group_id))
        conn.commit()
        log_action(session.get('username', 'невідомо'), f"змінив роль/групи користувача ID {user_id}")
        conn.close()
        return redirect(url_for('admin.user_list'))

    user = conn.execute("SELECT id, username, role FROM users WHERE id = ?", (user_id,)).fetchone()
    current_groups = conn.execute("SELECT group_id FROM user_groups WHERE user_id = ?", (user_id,)).fetchall()
    current_group_ids = [row['group_id'] for row in current_groups]
    groups = conn.execute("""
        SELECT id, name, start_year, study_form, program_credits,
               name || ' (' || start_year || ', ' || study_form || ', ' || program_credits || ' кредитів)' AS display_name
        FROM groups
        ORDER BY name, start_year
    """).fetchall()
    conn.close()

    log_action(session.get('username', 'невідомо'), f"відкрив форму редагування користувача ID {user_id}")
    return render_template('edit_user.html', user=user, groups=groups, current_group_ids=current_group_ids)

@admin_bp.route('/admin/group_export', methods=['GET', 'POST'])
@login_required('admin')
def group_export():
    """Отображение формы экспорта документов для группы."""
    conn = get_db()
    conn.row_factory = sqlite3.Row  # Используем Row для доступа к данным по именам колонок

    # Получение списка активных групп
    groups = conn.execute("""
        SELECT id, name, start_year, study_form, program_credits,
               name || ' (' || start_year || ', ' || study_form || ', ' || program_credits || ' кредитів)' AS display_name
        FROM groups
        WHERE archived = FALSE
        ORDER BY name, start_year
    """).fetchall()

    current_year = datetime.now().year
    years = list(range(1980, current_year + 1))
    students = []
    selected_group_id = request.args.get('group_id', type=int) if request.method == 'GET' else request.form.get('group_id', type=int)
    selected_year = request.args.get('birth_year', type=int) if request.method == 'GET' else request.form.get('birth_year', type=int)
    selected_template = request.args.get('template', 'template_word/template_adddiplom.docx') if request.method == 'GET' else request.form.get('template', 'template_word/template_adddiplom.docx')

    # Проверка корректности выбранных параметров
    if selected_group_id:
        group_check = conn.execute("SELECT id FROM groups WHERE id = ? AND archived = FALSE", (selected_group_id,)).fetchone()
        if not group_check:
            flash('Обрана група не існує або є архівною.', 'error')
            selected_group_id = None

    if request.method == 'POST':
        if not selected_group_id and not selected_year:
            flash('Будь ласка, оберіть групу або рік народження.', 'error')
        else:
            # Получаем список активных студентов из формы
            active_students = request.form.getlist('active_students')
            # Перенаправляем с параметрами, включая активных студентов
            return redirect(url_for('admin.generate_group_docs', group_id=selected_group_id, birth_year=selected_year, template=selected_template, active_students=','.join(active_students)))

    if selected_group_id or selected_year:
        base_query = """
            SELECT * FROM students
            WHERE archived = FALSE -- Исключение архивных студентов
        """
        params = []
        if selected_group_id:
            base_query += " AND group_id = ?"
            params.append(selected_group_id)
        if selected_year:
            base_query += " AND SUBSTR(birth_date, 7, 4) >= ?"
            params.append(str(selected_year))
        try:
            students = conn.execute(base_query, params).fetchall()
        except Exception as e:
            logging.error(f"Помилка при отриманні студентів: {e}, Запит: {base_query}, Параметри: {params}")
            conn.close()
            return "Помилка бази даних", 500

    conn.close()
    log_action(session.get('username', 'невідомо'), f"відкрив форму експорту документів для групи (group_id={selected_group_id}, birth_year={selected_year})")
    return render_template(
        'group_export.html',
        students=students,
        groups=groups,
        years=years,
        selected_group_id=selected_group_id,
        selected_year=selected_year,
        selected_template=selected_template
    )
    
# Папка для временного хранения загруженных файлов
UPLOAD_FOLDER = 'Uploads'
ALLOWED_EXTENSIONS = {'xlsx'}

# Проверка расширения файла
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@admin_bp.route('/admin/import_subjects', methods=['GET', 'POST'])
@login_required('admin')
def import_subjects():
    """Импорт предметов из Excel-файла."""
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT id, name, start_year, study_form, program_credits,
                   name || ' (' || start_year || ', ' || study_form || ', ' || program_credits || ' кредитів)' AS display_name
            FROM groups
            WHERE archived = FALSE
            ORDER BY name, start_year
        """)
        groups = cursor.fetchall()
        logging.info(f"Fetched {len(groups)} groups from database")
        if not groups:
            flash("Немає доступних груп", "warning")
    except sqlite3.Error as e:
        logging.error(f"Database error while fetching groups: {e}")
        flash("Помилка бази даних при отриманні груп", "error")
        groups = []

    selected_group_id = request.args.get('group_id', '')

    if request.method == 'POST':
        file = request.files.get('excel_file')
        group_id = request.form.get('group_id')

        if not file or not allowed_file(file.filename):
            flash("Будь ласка, виберіть файл формату .xlsx", "error")
            return render_template('import_subjects.html', groups=groups, selected_group_id=selected_group_id)
        
        if not group_id:
            flash("ID групи не вказано", "error")
            return render_template('import_subjects.html', groups=groups, selected_group_id=selected_group_id)
        
        try:
            group_id = int(group_id)
            cursor.execute('SELECT id FROM groups WHERE id = ?', (group_id,))
            if not cursor.fetchone():
                flash("Обрана група не існує", "error")
                return render_template('import_subjects.html', groups=groups, selected_group_id=selected_group_id)
        except ValueError:
            flash("Некоректний ID групи", "error")
            return render_template('import_subjects.html', groups=groups, selected_group_id=selected_group_id)

        filename = secure_filename(file.filename)
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        file.save(filepath)

        try:
            wb = openpyxl.load_workbook(filepath)
            sheet = wb.active

            inserted = 0
            skipped = 0

            if sheet.max_row < 2:
                flash("Excel-файл порожній або не містить даних", "error")
                os.remove(filepath)
                return render_template('import_subjects.html', groups=groups, selected_group_id=selected_group_id)

            cursor.execute("SELECT MAX(position) FROM subjects WHERE group_id = ?", (group_id,))
            max_position = cursor.fetchone()[0] or 0
            current_position = max_position + 1

            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    code, name, credits, type_ = row
                    if not all([code, name, credits, type_]):
                        flash(f"❗ Неповні дані у рядку {i}", "error")
                        skipped += 1
                        continue
                    if type_ not in ['Залік', 'Екзамен']:
                        flash(f"❗ Невірний тип у рядку {i}: {type_}", "error")
                        skipped += 1
                        continue
                    credits = int(credits)
                    if credits < 1:
                        flash(f"❗ Некоректні кредити у рядку {i}", "error")
                        skipped += 1
                        continue

                    cursor.execute("SELECT id FROM subjects WHERE group_id = ? AND code = ?", (group_id, code))
                    if cursor.fetchone():
                        flash(f"❗ Предмет з кодом {code} уже існує у групі {group_id}", "error")
                        skipped += 1
                        continue

                    cursor.execute(
                        "INSERT INTO subjects (code, name, credits, type, position, group_id) VALUES (?, ?, ?, ?, ?, ?)",
                        (code, name, credits, type_, current_position, group_id)
                    )
                    inserted += 1
                    current_position += 1
                except Exception as e:
                    flash(f"⚠️ Помилка в рядку {i}: {e}", "error")
                    skipped += 1
                    continue

            conn.commit()
            flash(f"✅ Імпорт завершено. Додано: {inserted}, пропущено: {skipped}", "success")
            logging.info(f"User {session.get('username', 'невідомо')} imported subjects: added {inserted}, skipped {skipped} for group_id={group_id}")
        except Exception as e:
            conn.rollback()
            flash(f"⚠️ Помилка при імпорті Excel: {e}", "error")
            logging.error(f"Error importing Excel for group_id={group_id}, subjects: {e}")
        finally:
            if os.path.exists(filepath):
                os.remove(filepath)
            conn.close()

        return redirect(url_for('admin.manage_subjects', group_id=group_id))

    logging.info(f"User {session.get('username', 'невідомо')} opened import_subjects form")
    conn.close()
    return render_template('import_subjects.html', groups=groups, selected_group_id=selected_group_id)

@admin_bp.route('/admin/generate_group_docs', methods=['GET', 'POST'])
@login_required('admin')
def generate_group_docs():
    """Генерация ZIP-архива с документами для группы."""
    group_id = request.args.get('group_id', type=int) if request.method == 'GET' else request.form.get('group_id', type=int)
    birth_year = request.args.get('birth_year', type=int) if request.method == 'GET' else request.form.get('birth_year', type=int)
    selected_template = request.args.get('template', 'template_word/template_adddiplom.docx') if request.method == 'GET' else request.form.get('template', 'template_word/template_adddiplom.docx')
    active_students = request.args.get('active_students', '').split(',') if request.args.get('active_students') else []

    # Проверка, что хотя бы один параметр фильтрации задан
    if not group_id and not birth_year:
        flash('Оберіть групу або рік народження для генерації документів.', 'error')
        return redirect(url_for('admin.group_export'))

    conn = get_db()
    conn.row_factory = sqlite3.Row
    base_query = """
        SELECT s.*, 
               g.name || ' (' || g.start_year || ', ' || g.study_form || ', ' || g.program_credits || ' кредитів)' AS group_name,
               g.study_form,
               g.start_year,
               g.program_credits,
               g.qualification_name,
               g.degree_level,
               g.specialty,
               g.educational_program,
               g.knowledge_area,
               g.qualification_name_en,
               g.degree_level_en,
               g.specialty_en,
               g.educational_program_en,
               g.knowledge_area_en,
               g.institution_name_and_status,
               g.institution_name_and_status_en,
               g.entry_requirements,
               g.entry_requirements_en,
               g.learning_outcomes,
               g.learning_outcomes_en,
               g.program_includes,
               g.program_includes_en
        FROM students s
        LEFT JOIN groups g ON s.group_id = g.id
        WHERE s.archived = FALSE  -- Исключение архивных студентов
    """
    params = []

    # Фильтр по группе, если указан
    if group_id:
        base_query += " AND s.group_id = ?"
        params.append(group_id)
    # Фильтр по году рождения, если указан
    if birth_year:
        base_query += " AND SUBSTR(s.birth_date, 7, 4) >= ?"
        params.append(str(birth_year))

    try:
        students = conn.execute(base_query, params).fetchall()
        if not students:
            logging.error(f"Студенты не найдены для group_id={group_id}, birth_year={birth_year}")
            conn.close()
            return "Студенты не найдены по заданным фильтрам", 404
    except Exception as e:
        logging.error(f"Ошибка при выполнении SQL-запроса: {e}")
        conn.close()
        return "Ошибка базы данных", 500

    # Фильтрация студентов по активным чекбоксам
    if active_students and active_students[0]:  # Проверяем, не пустой ли список
        students = [s for s in students if str(s['id']) in active_students]

    # Получение имени группы, если group_id задан
    group_name = "Зі всіх груп"
    if group_id and students:
        group_name = students[0]['group_name'] if students[0]['group_name'] else f"Група_{group_id}"

    output_dir = os.path.join(os.getcwd(), 'generated_docs')
    os.makedirs(output_dir, exist_ok=True)

    zip_filename = f"{group_name}_{str(birth_year) if birth_year else 'Всі роки народження'}.zip"
    zip_path = os.path.join(output_dir, zip_filename)

    with ZipFile(zip_path, 'w') as zipf:
        for student in students:
            student_dict = dict(student)
            military = conn.execute("SELECT * FROM military WHERE student_id = ?", (student['id'],)).fetchone()
            military_dict = dict(military) if military else {}

            filename = f"{student_dict['last_name_UA']}_{student_dict['first_name_UA']}.docx".replace(" ", "_")
            full_path = os.path.join(output_dir, filename)

            try:
                gen_doc(student_dict, military_dict, template=selected_template, out=full_path, user_name=session.get('username', 'невідомо'))
                zipf.write(full_path, arcname=filename)
            except Exception as e:
                logging.error(f"Ошибка при генерации документа для студента {student_dict.get('last_name_UA', '')}: {e}")
                continue

    log_action(session.get('username', 'невідомо'), f"згенерував документи для групи {group_name} (рік народження: {birth_year or 'всі'})")
    conn.close()
    return send_file(zip_path, as_attachment=True)
    
@admin_bp.route('/admin/archive/<int:group_id>', methods=['POST'])
@login_required('admin')
def archive_group(group_id):
    conn = get_db()
    cursor = conn.cursor()
    
    # Проверка существования группы
    cursor.execute("SELECT id FROM groups WHERE id = ?", (group_id,))
    if not cursor.fetchone():
        flash('Група не знайдена', 'error')
        conn.close()
        return redirect(url_for('admin.manage_groups'))
    
    # Маркировка группы как архивной
    cursor.execute("UPDATE groups SET archived = TRUE WHERE id = ?", (group_id,))
    
    # Маркировка студентов группы как архивных (опционально)
    cursor.execute("UPDATE students SET archived = TRUE WHERE group_id = ?", (group_id,))
    
    conn.commit()
    conn.close()
    
    log_action(session.get('username', 'невідомо'), f"заархівував групу ID {group_id}")
    flash('Групу успішно заархівовано', 'success')
    return redirect(url_for('admin.manage_groups'))
    
@admin_bp.route('/admin/unarchive_group/<int:group_id>', methods=['POST'])
@login_required('admin')
def unarchive_group(group_id):
    """Разархивирование группы и связанных студентов."""
    conn = get_db()
    cursor = conn.cursor()
    
    # Проверка существования группы
    cursor.execute("SELECT id FROM groups WHERE id = ? AND archived = TRUE", (group_id,))
    if not cursor.fetchone():
        flash('Архівна група не знайдена', 'error')
        conn.close()
        return redirect(url_for('admin.archive'))
    
    # Разархивирование группы
    cursor.execute("UPDATE groups SET archived = FALSE WHERE id = ?", (group_id,))
    
    # Разархивирование студентов группы
    cursor.execute("UPDATE students SET archived = FALSE WHERE group_id = ?", (group_id,))
    
    conn.commit()
    conn.close()
    
    log_action(session.get('username', 'невідомо'), f"розархівував групу ID {group_id}")
    flash('Групу успішно розархівовано', 'success')
    return redirect(url_for('admin.archive'))
   
@admin_bp.route('/admin/archive')
@login_required('admin')
def archive():
    """Отображение списка всех архивных групп с возможностью просмотра студентов."""
    conn = get_db()
    conn.row_factory = sqlite3.Row

    # Получение списка всех архивных групп
    groups = conn.execute("""
        SELECT g.id, g.name, g.start_year, g.study_form, g.program_credits,
               g.name || ' (' || g.start_year || ', ' || g.study_form || ', ' || g.program_credits || ' кредитів)' AS display_name,
               (SELECT COUNT(*) FROM students s WHERE s.group_id = g.id AND s.archived = TRUE) AS student_count
        FROM groups g
        WHERE g.archived = TRUE
        ORDER BY g.start_year DESC, g.name
    """).fetchall()

    # Инициализация данных студентов (можно загрузить по требованию через AJAX, но для простоты загружаем все сразу)
    students_by_group = {}
    for group in groups:
        students = conn.execute("""
            SELECT id, last_name_UA, first_name_UA, birth_date
            FROM students
            WHERE group_id = ? AND archived = TRUE
            ORDER BY last_name_UA
        """, (group['id'],)).fetchall()
        students_by_group[group['id']] = students

    conn.close()
    log_action(session.get('username', 'невідомо'), "переглянув список архівних груп")
    return render_template('archive.html', groups=groups, students_by_group=students_by_group)